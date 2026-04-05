#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Add latitude/longitude columns to K&B Moscow shops Excel via Nominatim."""

import json
import os
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path

import certifi

os.environ.setdefault("SSL_CERT_FILE", certifi.where())
os.environ.setdefault("REQUESTS_CA_BUNDLE", certifi.where())

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

WORKSPACE = Path(__file__).resolve().parent
XLSX_NAME = "Адреса_магазинов_Красное_Белое.xlsx"
CACHE_NAME = "geocode_cache_moscow.json"

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load_cache(path: Path) -> dict[str, dict[str, float | None]]:
    if not path.is_file():
        return {}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return {}
    out: dict[str, dict[str, float | None]] = {}
    for k, v in data.items():
        if isinstance(v, dict) and "lat" in v and "lon" in v:
            out[k] = {"lat": v.get("lat"), "lon": v.get("lon")}
    return out


def save_cache(path: Path, cache: dict[str, dict[str, float | None]]) -> None:
    path.write_text(
        json.dumps(cache, ensure_ascii=False, indent=0, sort_keys=True),
        encoding="utf-8",
    )


def norm_key(addr: str) -> str:
    return " ".join(addr.split()).strip().lower()


def geocode_photon(address: str) -> tuple[float | None, float | None]:
    """Геокодер Photon (OSM) — хорошо по адресам в Москве, без квоты Nominatim."""
    import ssl

    queries = [
        f"{address}, Москва, Россия",
        f"Москва, {address}",
        f"{address}, Москва",
    ]
    ctx = ssl.create_default_context(cafile=certifi.where())
    for q in queries:
        url = "https://photon.komoot.io/api/?" + urllib.parse.urlencode({"q": q, "limit": 1})
        try:
            with urllib.request.urlopen(url, timeout=20, context=ctx) as resp:
                data = json.loads(resp.read().decode())
        except (urllib.error.URLError, OSError, json.JSONDecodeError, ValueError, KeyError):
            time.sleep(0.5)
            continue
        feats = data.get("features") or []
        if not feats:
            continue
        geom = feats[0].get("geometry") or {}
        coords = geom.get("coordinates")
        if not coords or len(coords) < 2:
            continue
        lon, lat = float(coords[0]), float(coords[1])
        # МО и ТиНАО (в т.ч. Зеленоград, Троицк)
        if 54.8 < lat < 57.2 and 35.5 < lon < 40.2:
            return lat, lon
    return None, None


def geocode_one(_geocode, address: str) -> tuple[float | None, float | None]:
    time.sleep(0.35)
    return geocode_photon(address)


def style_coord_cell(cell) -> None:
    cell.border = BORDER
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    cell.font = Font(size=11)


def apply_cache_to_sheet(ws, cache: dict, addr_col: int, lat_col: int, lon_col: int, max_row: int) -> int:
    """Заполняет координаты из кэша без сети. Возвращает число обновлённых строк."""
    updated = 0
    for r in range(2, max_row + 1):
        addr = ws.cell(row=r, column=addr_col).value
        if not addr or not str(addr).strip():
            continue
        key = norm_key(str(addr).strip())
        if key not in cache or cache[key].get("lat") is None:
            continue
        lat_f, lon_f = cache[key]["lat"], cache[key]["lon"]
        for col, val in ((lat_col, lat_f), (lon_col, lon_f)):
            cell = ws.cell(row=r, column=col, value=val)
            style_coord_cell(cell)
        updated += 1
    return updated


def main() -> None:
    try:
        sys.stdout.reconfigure(line_buffering=True)  # type: ignore[attr-defined]
    except (AttributeError, OSError):
        pass
    xlsx_path = WORKSPACE / XLSX_NAME
    cache_path = WORKSPACE / CACHE_NAME
    if not xlsx_path.is_file():
        print(f"Не найден файл: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    cache = load_cache(cache_path)

    wb = load_workbook(xlsx_path)
    ws = wb.active

    # столбцы с адресом — 3 (C)
    addr_col = 3
    lat_col, lon_col = 6, 7

    if ws.cell(row=1, column=lat_col).value is None:
        ws.cell(row=1, column=lat_col, value="Широта (WGS84)")
    if ws.cell(row=1, column=lon_col).value is None:
        ws.cell(row=1, column=lon_col, value="Долгота (WGS84)")
    for c in (lat_col, lon_col):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    max_row = ws.max_row
    n = apply_cache_to_sheet(ws, cache, addr_col, lat_col, lon_col, max_row)
    if n:
        print(f"Из кэша подставлено строк (без сети): {n}", flush=True)
        wb.save(xlsx_path)

    done = 0
    failed: list[int] = []

    for r in range(2, max_row + 1):
        addr = ws.cell(row=r, column=addr_col).value
        if not addr or not str(addr).strip():
            continue
        addr_s = str(addr).strip()
        key = norm_key(addr_s)

        lat = ws.cell(row=r, column=lat_col).value
        lon = ws.cell(row=r, column=lon_col).value
        if lat is not None and lon is not None and str(lat).strip() != "" and str(lon).strip() != "":
            continue

        if key in cache:
            lat_f, lon_f = cache[key]["lat"], cache[key]["lon"]
        else:
            lat_f, lon_f = geocode_one(None, addr_s)
            cache[key] = {"lat": lat_f, "lon": lon_f}
            save_cache(cache_path, cache)
            done += 1
            if done % 10 == 0:
                wb.save(xlsx_path)
                print(f"… строка {r}/{max_row}, новых запросов к API: {done}", flush=True)
        if lat_f is None:
            failed.append(r)

        for col, val in ((lat_col, lat_f), (lon_col, lon_f)):
            cell = ws.cell(row=r, column=col, value=val)
            style_coord_cell(cell)

    ws.auto_filter.ref = f"A1:{get_column_letter(lon_col)}{max_row}"
    wb.save(xlsx_path)
    print(f"Готово. Строк данных: {max_row - 1}. Без координат: {len(failed)}", flush=True)
    if failed and len(failed) <= 30:
        print("Строки без координат:", failed)


if __name__ == "__main__":
    main()
